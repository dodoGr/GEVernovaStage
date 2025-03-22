import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from openpyxl.styles import Border, Side, Alignment

file_path = 'Exercice.xlsx' 

# Charger les feuilles avec pandas en utilisant openpyxl
Summary = pd.read_excel(file_path, sheet_name='Summary', engine='openpyxl')
README = pd.read_excel(file_path, sheet_name='Readme', engine='openpyxl')

# Charger le fichier Excel avec openpyxl et data_only=True pour obtenir les valeurs finales
wb = load_workbook(file_path, data_only=True)

def formatage(feuille, min_row, max_row, min_col, max_col):

    # Définir les styles de bordure
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Définir l'alignement au centre
    center_alignment = Alignment(horizontal='center', vertical='center')

    # Appliquer les bordures et l'alignement aux cellules
    for row in feuille.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_alignment

    # Fusionner les cellules (exemple)
    feuille.merge_cells("A2:B2")
    feuille.merge_cells("A3:A4")
    print(f"Formatage appliqué à la feuille '{feuille.title}'.")


def table(valIndex, name_page, name_title, name_test):
    with pd.ExcelWriter('Exercice.xlsx', engine='openpyxl', mode='a') as writer:
        for i in range(0, valIndex):  # Boucle pour créer autant de page que de colonnes sous 'SHEET'
            new_sheet_name = f'{name_page[i]}'

            modele = pd.DataFrame({
                'Colonne 1': [name_title[i], name_test[i], np.nan],
                'Colonne 2': [np.nan, name_test[i] + ".1", name_test[i] + ".2"]
    
            #    'Colonne 1': ["Case " + str(i+1), "Test " + str(i+1), np.nan],
            #    'Colonne 2': [np.nan, "Test " + str(i+1) + ".1", "Test " + str(i+1) + ".2"]
                
            })

            if new_sheet_name not in writer.book.sheetnames:
                # Ajouter une nouvelle feuille
                modele.to_excel(writer, sheet_name=new_sheet_name, index=False)
                print(f"Nouvelle feuille '{new_sheet_name}' ajoutée avec succès.")
            else:
                print(f"La feuille '{new_sheet_name}' existe déjà.")

    # Charger le fichier Excel avec openpyxl
    wb = load_workbook(file_path)
    for i in range(0, valIndex):  # appliquer le formatage à toutes les nouvelles pages
        new_sheet_name = f'{name_page[i]}'
        if new_sheet_name in wb.sheetnames:
            nouvellePage = wb[new_sheet_name]  # accéder à la page
            formatage(nouvellePage, min_row=1, max_row=4, min_col=1, max_col=2)

    # Sauvegarder les modifications
    wb.save(file_path)

def lire_colonne(feuille, colonne): #fonction qui permet de lire les valeurs d'un colonne tant quon ne trouve pas de cellule vide

    ligne = 3  
    valeurs = []  # liste pour stocker les valeurs lues

    while True:
        # Lire la valeur de la cellule
        valeur = feuille.cell(row=ligne, column=colonne).value

        # Arrêter la boucle si la cellule est vide
        if valeur is None:
            break

        # Ajouter la valeur à la liste
        valeurs.append(valeur)

        # Passer à la ligne suivante
        ligne += 1

    return valeurs


if __name__ == '__main__':

    # afficher dans le terminal le contenu des pages readme et summaary 
    print(README)
    print("=" * 75)
    print(Summary)
    print("=" * 75)

    sheet = wb['Summary']

    # Recherche de "SHEET" dans toutes les cellules
    search_for_sheet = Summary.isin(['SHEET'])  # Recherche en majuscules
    if search_for_sheet.any().any():  # Vérifie si "SHEET" est trouvé 
        indicesPage = search_for_sheet.stack()[lambda x: x].index[0]
        ligne, colonne = indicesPage  # Récupérer la ligne et la colonne
        print(f"Le mot 'SHEET' a été trouvé à la ligne {ligne+1}, colonne {colonne}")   
    else:
        print("Le mot 'SHEET' n'a pas été trouvé dans la feuille 'Summary'.")

    colonne_a_lire = Summary.columns.get_loc(colonne)  # Numéro de la colonne à lire
    valeurs_pages_lues = lire_colonne(sheet, colonne_a_lire+1)
    valeurs_titles_lues = lire_colonne(sheet, colonne_a_lire+2)
    valeurs_tests_lues = lire_colonne(sheet, colonne_a_lire+3)


    table(len(valeurs_pages_lues) ,valeurs_pages_lues, valeurs_titles_lues,valeurs_tests_lues)

    # Afficher les valeurs lues
    print(f"Valeurs lues dans la colonne {colonne_a_lire} : {valeurs_pages_lues}")
    print(f"Nombre de valeurs dans la colonne {colonne_a_lire} : {len(valeurs_pages_lues)}")

    print(Summary.columns)

    sheet.merge_cells("A3:A4")

    # Ouvrir le fichier Excel automatiquement
    if os.path.exists(file_path):
        os.startfile(file_path)
        wb.save(file_path)

    else:
        print(f"Le fichier '{file_path}' n'existe pas.")
